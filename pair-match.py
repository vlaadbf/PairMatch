# -*- coding: utf-8 -*-
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =========================
#   CONFIG & PATHS
# =========================
BASE = r''
PATH_MASCULI = os.path.join(BASE, '.csv')
PATH_FEMELE  = os.path.join(BASE, '.csv')

OUT_POTRIVIRE = os.path.join(BASE, 'POTRIVIRE.csv')
OUT_COMBINARE = os.path.join(BASE, 'COMBINARE_CULOARE_FEMELE.csv')
DIR_XLSX      = os.path.join(BASE, 'rezultat')
os.makedirs(DIR_XLSX, exist_ok=True)

# Culori (ARGB) — 1=albastru, 2=portocaliu/auriu, 3=verde
FILL_MAP = {
    1: PatternFill(fill_type='solid', start_color='FF00B0F0', end_color='FF00B0F0'),  # albastru
    2: PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000'),  # portocaliu/auriu
    3: PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050'),  # verde
    # 4 = fără culoare
}

def sanitize_filename(name: str) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        name = "NECUNOSCUT"
    name = str(name).strip()
    name = re.sub(r'[\\/*?:"<>|]', '_', name)
    return name if name else "NECUNOSCUT"

def colorize_and_cleanup(xlsx_path: str, targets, sources, remove_sources=True):
    """
    Colorează target-urile pe baza codurilor din sources (1=albastru, 2=portocaliu, 3=verde).
    Apoi șterge coloanele-sursă dacă remove_sources=True.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_row = 1
    header_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
    tgt_idx = [header_map.get(h) for h in targets]
    src_idx = [header_map.get(h) for h in sources]

    for r in range(2, ws.max_row + 1):
        for ti, si in zip(tgt_idx, src_idx):
            if ti is None or si is None:
                continue
            val = ws.cell(row=r, column=si).value
            try:
                code = int(val) if val is not None else None
            except (TypeError, ValueError):
                code = None
            ws.cell(row=r, column=ti).fill = FILL_MAP.get(code, PatternFill(fill_type=None))

    if remove_sources:
        header_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
        to_delete = [header_map.get(h) for h in sources if header_map.get(h)]
        for col_idx in sorted(to_delete, reverse=True):
            ws.delete_cols(col_idx, 1)

    wb.save(xlsx_path)
    wb.close()

# -------------------------
# Utilitare normalizare
# -------------------------
def normalize_diacritics(s: str) -> str:
    return (s.replace('Ă', 'A').replace('Â', 'A').replace('Î', 'I')
             .replace('Ş', 'S').replace('Ș', 'S').replace('Ţ', 'T').replace('Ț', 'T'))

def norm_text(x):
    if pd.isna(x):
        return ''
    return normalize_diacritics(str(x).strip().upper())

def norm_rasa(x: str) -> str:
    return norm_text(x)

def norm_sectiune(x: str) -> str:
    s = norm_text(x)
    s = s.replace('SUPLIMENTARĂ', 'SUPLIMENTARA')
    return s if s else 'PRINCIPALA'

# =========================
#   PASUL 0: CITIRE + NORMALIZĂRI
# =========================
masculi_df = pd.read_csv(PATH_MASCULI, dtype=str)
femele_df  = pd.read_csv(PATH_FEMELE,  dtype=str)

# curățare headere (spații)
masculi_df.columns = masculi_df.columns.str.strip()
femele_df.columns  = femele_df.columns.str.strip()

# Conversie A/P/V -> 1/2/3 (MASCULI)
if 'CULOARE' in masculi_df.columns:
    color_map = {'A': 1, 'P': 2, 'V': 3, 'a': 1, 'p': 2, 'v': 3, '1': 1, '2': 2, '3': 3}
    masculi_df['CULOARE'] = masculi_df['CULOARE'].map(color_map).fillna(4)
    masculi_df['CULOARE'] = pd.to_numeric(masculi_df['CULOARE'], errors='coerce').fillna(4).astype(int)

# Conversie A/P/V -> 1/2/3 (FEMELE)
if 'CULOARE' in femele_df.columns:
    color_map = {'A': 1, 'P': 2, 'V': 3, 'a': 1, 'p': 2, 'v': 3, '1': 1, '2': 2, '3': 3}
    femele_df['CULOARE'] = femele_df['CULOARE'].map(color_map).fillna(4)
    femele_df['CULOARE'] = pd.to_numeric(femele_df['CULOARE'], errors='coerce').fillna(4).astype(int)

# Normalizări RASA + SECTIUNE
masculi_df['RASA_NORM'] = masculi_df['RASA'].apply(norm_rasa) if 'RASA' in masculi_df.columns else ''
femele_df['RASA_NORM']  = femele_df['RASA'].apply(norm_rasa)  if 'RASA' in femele_df.columns else ''

if 'SECTIUNE' in femele_df.columns:
    femele_df['SECTIUNE'] = femele_df['SECTIUNE'].apply(norm_sectiune)
else:
    femele_df['SECTIUNE'] = 'PRINCIPALA'

sect_rank = {'PRINCIPALA': 0, 'SUPLIMENTARA 1': 1, 'SUPLIMENTARA 2': 2}
femele_df['SECTIUNE_RANK'] = femele_df['SECTIUNE'].map(sect_rank).fillna(3).astype(int)

# >>> NOU: RASA_RANK — non-METIS (0) înainte de METIS (1) pentru sortare
femele_df['RASA_RANK'] = (femele_df['RASA_NORM'].eq('METIS')).astype(int)

# --- Regula METIS: identificare crescători eligibili ---
# crescători cu EXACT 1 rasă non-METIS în femele (și cel puțin una non-METIS)
femele_non_metis = femele_df[femele_df['RASA_NORM'].ne('METIS')]
cnt_per_cresc = femele_non_metis.groupby('CRESCATOR')['RASA_NORM'].nunique(dropna=True)
eligible_breeders_for_metis = set(cnt_per_cresc[cnt_per_cresc == 1].index)

# =========================
#   PASUL 1: POTRIVIRE.csv
# =========================
cols_fem = [
    'CRESCATOR','Matricol','RASA','RASA_NORM','MAMA','TATA','JUDET','DATA NASTERE',
    'CULOARE','VA','CONTROLOR','SECTIUNE','SECTIUNE_RANK','RASA_RANK'
]
potrivire_all = pd.merge(
    masculi_df,
    femele_df[cols_fem],
    on='CRESCATOR',
    suffixes=('_Masculi', '_Femele')
)

# >>> Asigur sufixe pentru coloanele calculate doar în FEMELE
for col in ['SECTIUNE_RANK', 'RASA_NORM', 'RASA_RANK']:
    if col in potrivire_all.columns and f'{col}_Femele' not in potrivire_all.columns:
        potrivire_all = potrivire_all.rename(columns={col: f'{col}_Femele'})

# Filtre genetice
potrivire_all = potrivire_all[potrivire_all['Matricol_Masculi'] != potrivire_all['TATA_Femele']]
potrivire_all = potrivire_all[potrivire_all['MAMA_Masculi']    != potrivire_all['MAMA_Femele']]
potrivire_all = potrivire_all[potrivire_all['MAMA_Masculi']    != potrivire_all['Matricol_Femele']]
potrivire_all = potrivire_all[potrivire_all['TATA_Masculi']    != potrivire_all['TATA_Femele']]

# Regula pe RASĂ:
# - accept dacă RASA_NORM_Masculi == RASA_NORM_Femele
# - SAU dacă femela e METIS și crescătorul e eligibil (are o singură rasă non-METIS în femele)
if 'RASA_NORM_Masculi' not in potrivire_all.columns and 'RASA_NORM' in masculi_df.columns:
    potrivire_all = potrivire_all.rename(columns={'RASA_NORM': 'RASA_NORM_Masculi'})

cond_same_race = potrivire_all['RASA_NORM_Masculi'] == potrivire_all['RASA_NORM_Femele']
cond_metis_ok  = (potrivire_all['RASA_NORM_Femele'].eq('METIS') &
                  potrivire_all['CRESCATOR'].isin(eligible_breeders_for_metis))
potrivire_all = potrivire_all[cond_same_race | cond_metis_ok]

# Selectăm câmpurile necesare mai departe
potrivire_filtered = potrivire_all[[
    'CRESCATOR',
    'Matricol_Masculi','DATA NASTERE_Masculi','MAMA_Masculi','TATA_Masculi',
    'Matricol_Femele','DATA NASTERE_Femele','MAMA_Femele','TATA_Femele',
    'JUDET_Masculi','CULOARE_Masculi','CULOARE_Femele','VA_Femele','CONTROLOR_Masculi',
    'RASA_Femele','SECTIUNE_Femele','SECTIUNE_RANK_Femele','RASA_RANK_Femele'
]]

potrivire_filtered.to_csv(OUT_POTRIVIRE, index=False, encoding='utf-8-sig')
print(f"[OK] POTRIVIRE -> {OUT_POTRIVIRE}")

# =========================
#   PASUL 2: COMBINARE_CULOARE_FEMELE.csv
#   Priorități (pe fiecare CRESCATOR):
#   0) RASA_RANK_Femele: non-METIS (0) -> METIS (1)
#   1) CULOARE_Femele: 1 -> 2 -> 3
#   2) SECTIUNE_RANK_Femele: PRINCIPALA -> SUPLIMENTARA 1 -> SUPLIMENTARA 2
#   3) CULOARE_Masculi: 1 -> 2 -> 3
#   + Balansare: la egalitate pe criterii, aleg masculul cu cele mai puține combinații
#   + Soft cap 50, hard cap 60
#   + IMPORTANT: masculul poate apărea la mai mulți crescători -> limitare/BALANSARE pe (CRESCATOR, Matricol_Masculi)
# =========================
tmp = potrivire_filtered.copy()
for col in ['CULOARE_Masculi','CULOARE_Femele','VA_Femele','SECTIUNE_RANK_Femele','RASA_RANK_Femele']:
    if col in tmp.columns:
        tmp[col] = pd.to_numeric(tmp[col], errors='coerce')

# ordinea ta de prioritate (fără balansare aici)
tmp = tmp.sort_values(
    ['CRESCATOR', 'RASA_RANK_Femele', 'CULOARE_Femele', 'SECTIUNE_RANK_Femele', 'CULOARE_Masculi', 'Matricol_Femele'],
    ascending=[True, True, True, True, True, True]
)

SOFT_CAP = 50
HARD_CAP = 60

combinari = {}                 # femela -> row selectat
numar_combinari_masculi = {}   # (CRESCATOR, mascul) -> count

def _k(crescator, mascul):
    return (str(crescator).strip().upper(), str(mascul).strip().upper())

def male_count(crescator, mascul):
    return numar_combinari_masculi.get(_k(crescator, mascul), 0)

def can_use_male(crescator, mascul, soft_allowed: bool):
    c = male_count(crescator, mascul)
    if c >= HARD_CAP:
        return False
    if (not soft_allowed) and c >= SOFT_CAP:
        return False
    return True

# procesăm femelele în ordinea dată de tmp
femele_order = tmp['Matricol_Femele'].dropna().drop_duplicates().tolist()

for femela in femele_order:
    cand = tmp[tmp['Matricol_Femele'] == femela]
    if cand.empty:
        continue

    cresc = cand['CRESCATOR'].iloc[0]  # femela aparține unui singur crescător

    # definim "cheia" de prioritate (fără balansare)
    cand = cand.copy()
    cand['PRIO_KEY'] = list(zip(
        cand['RASA_RANK_Femele'].fillna(999).astype(int),
        cand['CULOARE_Femele'].fillna(999).astype(int),
        cand['SECTIUNE_RANK_Femele'].fillna(999).astype(int),
        cand['CULOARE_Masculi'].fillna(999).astype(int),
    ))

    best_row = None

    # chei de prioritate în ordine crescătoare
    for prio in sorted(cand['PRIO_KEY'].unique()):
        block = cand[cand['PRIO_KEY'] == prio]

        # soft cap per (crescator, mascul)
        any_under_soft = any(
            male_count(cresc, m) < SOFT_CAP
            for m in block['Matricol_Masculi'].tolist()
        )
        soft_allowed = not any_under_soft  # dacă NU există alternativă sub soft, permitem peste soft

        # sortăm candidații din acest block după:
        # 1) count mascul (balansare) per crescător
        # 2) tie-break stabil (Matricol_Masculi)
        block = block.copy()
        block['MALE_COUNT'] = block['Matricol_Masculi'].apply(lambda m: male_count(cresc, m)).astype(int)
        block = block.sort_values(['MALE_COUNT', 'Matricol_Masculi'], ascending=[True, True])

        for _, row in block.iterrows():
            m = row['Matricol_Masculi']
            if can_use_male(cresc, m, soft_allowed=soft_allowed):
                best_row = row
                break

        if best_row is not None:
            break

    if best_row is None:
        continue

    combinari[femela] = best_row
    cresc2 = best_row['CRESCATOR']
    m2 = best_row['Matricol_Masculi']
    numar_combinari_masculi[_k(cresc2, m2)] = male_count(cresc2, m2) + 1

combinari_df = pd.DataFrame(list(combinari.values()))
if not combinari_df.empty:
    # în final sortăm și după mascul (penultimul, înainte de Matricol_Femele)
    combinari_df = combinari_df.sort_values(
        [
            'CONTROLOR_Masculi',
            'CRESCATOR',
            'RASA_RANK_Femele',
            'CULOARE_Femele',
            'SECTIUNE_RANK_Femele',
            'CULOARE_Masculi',
            'Matricol_Masculi',
            'Matricol_Femele'
        ],
        ascending=[True, True, True, True, True, True, True, True]
    )

combinari_df.to_csv(OUT_COMBINARE, index=False, encoding='utf-8-sig')
print(f"[OK] COMBINARE_CULOARE_FEMELE -> {OUT_COMBINARE}")

# =========================
#   PASUL 3: Excel-uri per CONTROLOR (ordine finală + colorare)
# =========================
if combinari_df.empty:
    print("[INFO] Nu există combinații după filtrare; nu s-au generat Excel-uri.")
else:
    final_map = {
        'CRESCATOR': 'CRESCATOR',
        'Matricol_Femele': 'MATRICOL',
        'RASA_Femele': 'RASA',
        'VA_Femele': 'VA',
        'DATA NASTERE_Femele': 'DATA NASTERE',
        'SECTIUNE_Femele': 'SECTIUNE',
        'Matricol_Masculi': 'MASCUL',
    }
    order_src = list(final_map.keys())

    for controlor, grup in combinari_df.groupby('CONTROLOR_Masculi'):
        # sortare sigură în fiecare Excel, tot cu mascul înainte de Matricol_Femele
        grup = grup.sort_values(
            [
                'CRESCATOR',
                'RASA_RANK_Femele',
                'CULOARE_Femele',
                'SECTIUNE_RANK_Femele',
                'CULOARE_Masculi',
                'Matricol_Masculi',
                'Matricol_Femele'
            ],
            ascending=[True, True, True, True, True, True, True]
        )

        nume = sanitize_filename(controlor)
        out_xlsx = os.path.join(DIR_XLSX, f'POTRIVIRE_{nume}.xlsx')
        sheet_name = (nume or "FOAIE")[:31] if nume else "FOAIE"

        # exportăm doar coloanele finale + culorile (pentru colorare), în ordinea cerută
        cols_for_write = order_src + ['CULOARE_Masculi', 'CULOARE_Femele']
        df_export = grup[cols_for_write].rename(columns=final_map)

        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name=sheet_name)

        # colorăm MASCUL (după CULOARE_Masculi) și MATRICOL (după CULOARE_Femele), apoi ștergem CULOARE_*
        colorize_and_cleanup(
            out_xlsx,
            targets=['MASCUL', 'MATRICOL'],
            sources=['CULOARE_Masculi', 'CULOARE_Femele'],
            remove_sources=True
        )

        print(f"[OK] Excel controlor -> {out_xlsx}")

print(f"[DONE] Fișiere generate:\n - {OUT_POTRIVIRE}\n - {OUT_COMBINARE}\n - {DIR_XLSX}")